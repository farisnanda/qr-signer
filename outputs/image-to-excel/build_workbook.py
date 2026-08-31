import ast
import base64
from io import BytesIO
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

base = Path("H:/qr-signer/outputs/image-to-excel")
js_text = (base / "build_workbook.mjs").read_text(encoding="utf-8")
match = re.search(r"const rows = (\[[\s\S]*?\n\]);", js_text)
if not match:
    raise SystemExit("Data rows not found")

rows = ast.literal_eval(match.group(1).replace("null", "None"))
output = base / "py_write.txt"

wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.append(["NO", "NAMA", "JK", "TUJUAN PENEMPATAN"])
for row in rows:
    ws.append(row)

thin = Side(style="thin", color="7F7F7F")
header_fill = PatternFill("solid", fgColor="D9EAF7")
for row in ws.iter_rows(min_row=1, max_row=78, min_col=1, max_col=4):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=cell.column == 4)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in range(2, 79):
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 58
ws.row_dimensions[1].height = 26
for row in range(2, 79):
    ws.row_dimensions[row].height = 22
for row in range(72, 78):
    ws.row_dimensions[row].height = 34

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:D78"
table = Table(displayName="PenempatanTable", ref="A1:D78")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

dv = DataValidation(type="list", formula1='"L,P"', allow_blank=False)
ws.add_data_validation(dv)
dv.add("C2:C78")

info = wb.create_sheet("Info")
info.sheet_view.showGridLines = False
info["A1"] = "Keterangan"
info.merge_cells("A1:B1")
info["A1"].fill = header_fill
info["A1"].font = Font(name="Arial", size=12, bold=True)
info_rows = [
    ("Sumber", "Diketik ulang dari 3 foto WhatsApp yang diberikan pengguna."),
    ("Dokumen", "Lampiran Surat Menteri Dalam Negeri Nomor 800.1.2.3/5533/SJ"),
    ("Tanggal dokumen", "29 Juli 2026"),
    ("Judul", "Penempatan Lulusan IPDN Angkatan XXXIII Tahun 2026"),
    ("Provinsi", "Jawa Timur"),
]
for idx, values in enumerate(info_rows, start=2):
    info.cell(idx, 1, values[0])
    info.cell(idx, 2, values[1])
for row in info.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(name="Arial", size=10, bold=cell.column == 1 and cell.row > 1)
for row in range(2, 7):
    info.cell(row, 1).fill = PatternFill("solid", fgColor="F2F2F2")
info.column_dimensions["A"].width = 20
info.column_dimensions["B"].width = 70

buffer = BytesIO()
wb.save(buffer)
payload = buffer.getvalue()
print(base64.b64encode(payload).decode("ascii"))
raise SystemExit

check = load_workbook(output, data_only=False)
data = check["Data"]
assert data.max_row == 78
assert data.max_column == 4
assert data["A78"].value == 77
assert data["D78"].value == "Kementerian Pertanian"
print(output)
print(f"rows={data.max_row - 1}, cols={data.max_column}, bytes={output.stat().st_size}")
